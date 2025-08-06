#!/usr/bin/env python
"""
Generate Excel Report from Extracted PDF Entities
Based on the format shown in sample_output.jpg
"""

import os
import json
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from datetime import datetime

def load_entities(json_file):
    """Load extracted entities from JSON file"""
    with open(json_file, 'r') as f:
        return json.load(f)

def generate_excel_report(entities, template_path, output_file):
    """Generate Excel report based on extracted entities matching sample_output.jpg format"""
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lab Analysis Report"
    
    # Define styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    green_font = Font(color='008000', bold=True, size=10)  # Green for Yes
    red_font = Font(color='FF0000', bold=True, size=10)    # Red for No
    centered = Alignment(horizontal='center', vertical='center')
    left_aligned = Alignment(horizontal='left', vertical='center')
    right_aligned = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green background
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')    # Light red background
    
    # Title Section
    current_row = 1
    ws['A1'] = "Laboratory Certificate/Analysis Report"
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = centered
    
    current_row += 2
    
    # Basic Information Section
    basic_info = [
        ("Our Ref:", entities.get('our_ref', 'Not found')),
        ("Company Name:", entities.get('company_name', 'Not found')),
        ("Lab Report Creation Date:", entities.get('lab_report_creation_date', 'Not found')),
        ("Subject:", entities.get('subject', 'Not found')),
        ("Sample Reference:", entities.get('sample_reference', 'Not found')),
        ("Is There Signature?:", entities.get('is_there_signature', 'Not found')),
        ("Results Comply?:", entities.get('results_comply', 'Not found'))
    ]
    
    for label, value in basic_info:
        ws.cell(row=current_row, column=1).value = label
        ws.cell(row=current_row, column=1).font = subtitle_font
        ws.cell(row=current_row, column=1).alignment = left_aligned
        
        value_cell = ws.cell(row=current_row, column=2)
        value_cell.value = value
        ws.merge_cells(f'B{current_row}:E{current_row}')
        value_cell.alignment = left_aligned
        
        # Apply special formatting for signature and compliance fields
        if "Is There Signature?" in label or "Results Comply?" in label:
            if str(value).lower() == "yes":
                value_cell.font = green_font
                value_cell.fill = green_fill
            elif str(value).lower() == "no":
                value_cell.font = red_font
                value_cell.fill = red_fill
            else:
                value_cell.font = normal_font
        else:
            value_cell.font = normal_font
        
        current_row += 1
    
    current_row += 1
    
    # Test Results Section
    ws.cell(row=current_row, column=1).value = "TEST RESULTS"
    ws.cell(row=current_row, column=1).font = header_font
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws.cell(row=current_row, column=1).alignment = centered
    ws.cell(row=current_row, column=1).fill = header_fill
    
    # Test results table header
    current_row += 1
    header_row = current_row
    headers = ["Parameter", "Unit", "Test Method", "Result", "Specification"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = centered
        cell.border = thin_border
        cell.fill = header_fill
    
    # Test results data
    current_row += 1
    test_results = entities.get('test_results', [])
    
    for test in test_results:
        row = current_row
        
        # Parameter
        ws.cell(row=row, column=1).value = test.get('parameter', '')
        ws.cell(row=row, column=1).alignment = left_aligned
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).font = normal_font
        
        # Unit
        ws.cell(row=row, column=2).value = test.get('unit', '')
        ws.cell(row=row, column=2).alignment = centered
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).font = normal_font
        
        # Test Method
        ws.cell(row=row, column=3).value = test.get('test_method', '')
        ws.cell(row=row, column=3).alignment = centered
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=3).font = normal_font
        
        # Result
        ws.cell(row=row, column=4).value = test.get('result', '')
        ws.cell(row=row, column=4).alignment = centered
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).font = normal_font
        
        # Specification
        ws.cell(row=row, column=5).value = test.get('specification', '')
        ws.cell(row=row, column=5).alignment = centered
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=5).font = normal_font
        
        current_row += 1
    
    # Footer
    current_row += 2
    ws.cell(row=current_row, column=1).value = "Report Generated by PDF Extraction System"
    ws.cell(row=current_row, column=1).font = Font(italic=True, size=9)
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws.cell(row=current_row, column=1).alignment = centered
    
    current_row += 1
    ws.cell(row=current_row, column=1).value = f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws.cell(row=current_row, column=1).font = Font(italic=True, size=9)
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws.cell(row=current_row, column=1).alignment = centered
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    
    # Save the workbook
    wb.save(output_file)
    print(f"Excel report generated: {output_file}")
    return output_file

def main():
    """Main function"""
    if len(sys.argv) < 2:
        print("Usage: python generate_excel_report.py <entities_json_file> [output_excel_file]")
        return
    
    # Get the entities JSON file
    entities_json = sys.argv[1]
    if not os.path.exists(entities_json):
        print(f"Error: File not found - {entities_json}")
        return
    
    # Determine output file name
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    else:
        output_file = Path(entities_json).stem.replace("_entities", "") + "_report.xlsx"
    
    # Generate the report
    try:
        entities = load_entities(entities_json)
        generate_excel_report(entities, None, output_file)
    except Exception as e:
        print(f"Error generating Excel report: {str(e)}")

if __name__ == "__main__":
    main()
